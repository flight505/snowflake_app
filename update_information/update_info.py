import os
import openai
import openpyxl
import json

secrets = openai.api_key = os.environ["OPENAI_API_KEY"]

print("OpenAI API key: " + secrets)

# Load Excel data
workbook = openpyxl.load_workbook("/Users/jesper/Projects/MTX/MTX data 2023/src/data/ALL08_PatientData_2021-01-26.xlsx")
sheet = workbook["ListVariables"]
max_column = sheet.max_column

# Loop over columns starting from column 2
for column in range(2, max_column + 1):
    
    # Get column name and description
    column_name = sheet.cell(row=1, column=column).value
    description = sheet.cell(row=2, column=column).value

    # Check if column has data starting from row 6
    row = 6
    data = {}
    while sheet.cell(row=row, column=column).value is not None:
        key, value = sheet.cell(row=row, column=column).value.split(" | ")
        data[key] = value
        row += 1

    # If no data in column, skip to next column
    if not data:
        continue

    # Create JSON file
    json_data = {
        "column_name": column_name,
        "description": description,
        "details": data
    }
    with open(f"update_information/{column_name}.json", "w") as outfile:
        json.dump(json_data, outfile)

    # Send JSON file to OpenAI chatbot and update description
    for _ in range(1):
        with open(f"update_information/{column_name}.json", "r") as infile:
            data = json.load(infile)
            prompt = 'Follow these guidelines:\n'
            prompt += f'Please provide a medically concise description for the {column_name} column.\n'
            prompt += f'The initial description for this column is "{description}", and it relates to content of this column and childhood Leukemia cancer treatment.\n'
            prompt += f'Please provide a brief and accurate description that accurately based on the {column_name} and the inital {description} and column {data}.\n'
            prompt += 'Keep the description short one line only.\n'
            prompt += 'Please do not include any personal information in your description.\n'
            prompt += 'Please do not include any information that is not medically relevant.\n'
            prompt += 'Please do not include any information that is not related to the column.\n'
            prompt += 'Please remove newline from the description\n'
            prompt += 'Thank you for your help in improving this medical dataset.'
            response = openai.Completion.create(
                engine="text-davinci-003",
                prompt=prompt,
                max_tokens=100,
                n=2,
                stop=None,
                temperature=0.2,
            )
            new_description = response.choices[0].text.strip()
            data["description"] = new_description

            # Update JSON file
            json_data = json.dumps(data, indent=4)
            with open(f"update_information/updated_gpt/{column_name}.json", "w") as outfile:
                outfile.write(json_data)

            print(f"New description: {new_description}")

